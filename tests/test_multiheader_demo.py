from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from spreadsheet_handling.cli.apps.run import main as run_main


def test_product_multiheader_workbook_has_grouped_headers(tmp_path, monkeypatch) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    monkeypatch.chdir(repo_root)

    workbook_path = tmp_path / "product_multiheader.xlsx"
    exit_code = run_main([
        "--config",
        "pipelines/demo_product_multiheader.yaml",
        "--out-path",
        str(workbook_path),
    ])
    assert exit_code == 0

    wb = load_workbook(workbook_path)
    try:
        ws = wb["product_overview"]
        merged = {str(rng) for rng in ws.merged_cells.ranges}
        assert ws["A1"].value == "product"
        assert ws["A2"].value == "id"
        assert ws["B2"].value == "name"
        assert ws["C1"].value == "manager"
        assert ws["D1"].value == "branch"
        assert "A1:B1" in merged
        assert "D1:E1" in merged
    finally:
        wb.close()
