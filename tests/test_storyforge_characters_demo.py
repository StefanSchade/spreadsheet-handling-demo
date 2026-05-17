from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from spreadsheet_handling.cli.apps.run import main as run_main


def test_storyforge_first_slice_materializes_species_helper(tmp_path, monkeypatch) -> None:
    """Optional secondary trail: characters -> species FK helper, no matrix/reimport."""
    repo_root = Path(__file__).resolve().parents[1]
    monkeypatch.chdir(repo_root)

    workbook_path = tmp_path / "storyforge_characters.xlsx"
    exit_code = run_main([
        "--config",
        "pipelines/demo_storyforge_characters.yaml",
        "--out-path",
        str(workbook_path),
    ])
    assert exit_code == 0

    wb = load_workbook(workbook_path)
    try:
        ws = wb["characters"]
        headers = {str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)}
        first_row = {
            str(ws.cell(row=1, column=c).value): ws.cell(row=2, column=c).value
            for c in range(1, ws.max_column + 1)
        }
    finally:
        wb.close()

    assert "_species_name" in headers
    assert first_row["_species_name"] == "Emberfox"
