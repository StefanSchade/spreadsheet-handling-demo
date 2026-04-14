from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from spreadsheet_handling.cli.apps.run import main as run_main


def _run_demo(repo_root: Path, *argv: str) -> None:
    exit_code = run_main(list(argv))
    assert exit_code == 0, f"demo run failed from {repo_root}"


def _header_map(ws) -> dict[str, int]:
    return {str(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}


def test_product_roundtrip_editing_preserves_user_changes(tmp_path, monkeypatch) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    monkeypatch.chdir(repo_root)

    workbook_path = tmp_path / "product_business_slice.xlsx"
    reimport_dir = tmp_path / "reimported"

    _run_demo(
        repo_root,
        "--config",
        "pipelines/demo_product_business_slice.yaml",
        "--out-path",
        str(workbook_path),
    )

    wb = load_workbook(workbook_path)
    try:
        product_ws = wb["product"]
        branch_ws = wb["branch"]

        product_headers = _header_map(product_ws)
        branch_headers = _header_map(branch_ws)

        product_ws.cell(row=2, column=product_headers["status"]).value = "pilot"
        branch_ws.cell(row=2, column=branch_headers["region"]).value = "EMEA"

        wb.save(workbook_path)
    finally:
        wb.close()

    _run_demo(
        repo_root,
        "--config",
        "pipelines/demo_workbook_reimport.yaml",
        "--in-path",
        str(workbook_path),
        "--out-path",
        str(reimport_dir),
    )

    products = json.loads((reimport_dir / "product.json").read_text(encoding="utf-8"))
    branches = json.loads((reimport_dir / "branch.json").read_text(encoding="utf-8"))

    assert products[0]["status"] == "pilot"
    assert branches[0]["region"] == "EMEA"
